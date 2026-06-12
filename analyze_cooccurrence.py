import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ttp_iter import load_reports, iter_report_ttps


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a TTP co-occurrence spreadsheet from a JSON dataset",
    )
    parser.add_argument(
        "dataset",
        help="Path to the input dataset JSON file",
    )
    parser.add_argument(
        "-o",
        "--output",
        default=None,
        help="Path to the output .xlsx file. Defaults to <dataset stem>-cooccurrence.xlsx",
    )
    return parser.parse_args()


def normalize_id(value: str) -> str:
    return value.strip().upper()


def load_reports(dataset_path: Path) -> list[dict[str, object]]:
    with dataset_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise ValueError("Expected the dataset root to be a JSON array of reports.")
    return data


@dataclass
class CooccurrenceData:
    # pair_counts[a][b] = number of reports containing both TTP a and TTP b
    pair_counts: dict[str, dict[str, int]]
    # ttp_counts[a] = number of reports containing TTP a
    ttp_counts: dict[str, int]
    # tactic_ttp_counts[tactic][ttp] = number of reports containing at least one
    #   TTP from `tactic` AND also containing `ttp`
    tactic_ttp_counts: dict[str, dict[str, int]]
    # tactic_counts[tactic] = number of reports containing at least one TTP from tactic
    tactic_counts: dict[str, int]
    # total number of reports
    total_reports: int
    ttp_names: dict[str, str]
    # ttp_tactic[ttp_id] = tactic name (or None)
    ttp_tactic: dict[str, str | None]


def build_cooccurrence_data(reports: list[dict[str, object]]) -> CooccurrenceData:
    pair_counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    ttp_counts: dict[str, int] = defaultdict(int)
    tactic_ttp_counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    tactic_counts: dict[str, int] = defaultdict(int)
    ttp_names: dict[str, str] = {}
    ttp_tactic: dict[str, str | None] = {}

    for report in reports:
        records = list({r.id: r for r in iter_report_ttps(report, include_goal=True)}.values())
        ttp_ids: set[str] = set()
        tactics_in_report: set[str] = set()

        for record in records:
            ttp_names.setdefault(record.id, record.name)
            ttp_tactic.setdefault(record.id, record.tactic)
            ttp_ids.add(record.id)
            if record.tactic:
                tactics_in_report.add(record.tactic)

        for ttp_b in ttp_ids:
            ttp_counts[ttp_b] += 1
            for ttp_a in ttp_ids:
                pair_counts[ttp_a][ttp_b] += 1

        for tactic in tactics_in_report:
            tactic_counts[tactic] += 1
            for ttp_b in ttp_ids:
                tactic_ttp_counts[tactic][ttp_b] += 1

    return CooccurrenceData(
        pair_counts, ttp_counts, tactic_ttp_counts, tactic_counts,
        len(reports), ttp_names, ttp_tactic,
    )


def ordered_ttps(ttp_names: dict[str, str]) -> list[str]:
    return sorted(ttp_names, key=lambda ttp_id: (ttp_id, ttp_names[ttp_id].lower()))


def compute_individual_probability(data: CooccurrenceData, ttp_a: str, ttp_b: str) -> float:
    """P(ttp_a appears | ttp_b appears)"""
    denominator = data.ttp_counts.get(ttp_b, 0)
    if denominator == 0:
        return 0.0
    return data.pair_counts.get(ttp_a, {}).get(ttp_b, 0) / denominator


def compute_tactic_given_ttp_probability(
    data: CooccurrenceData, row_tactic: str, col_ttp: str
) -> float:
    """P(row_tactic present | col_ttp appears)"""
    denominator = data.ttp_counts.get(col_ttp, 0)
    if denominator == 0:
        return 0.0
    numerator = data.tactic_ttp_counts.get(row_tactic, {}).get(col_ttp, 0)
    return numerator / denominator


def prior_probability(data: CooccurrenceData, col_ttp: str) -> float:
    """P(col_ttp appears) = ttp_counts[col_ttp] / total_reports"""
    if data.total_reports == 0:
        return 0.0
    return data.ttp_counts.get(col_ttp, 0) / data.total_reports


def sheet_name_for_tactic(tactic: str) -> str:
    invalid = r"\/*?:[]{}"
    cleaned = "".join(c if c not in invalid else "_" for c in tactic)
    return cleaned[:31]


def write_matrix(
    sheet,
    row_labels: list[str],
    col_labels: list[str],
    value_getter,   # callable(row_label, col_label) -> float
    number_format: str,
    row_header: str,
) -> None:
    header_font = Font(bold=True)
    sheet.freeze_panes = "B2"
    sheet.append([row_header, *col_labels])
    for cell in sheet[1]:
        cell.font = header_font

    for row_label in row_labels:
        row = [row_label]
        for col_label in col_labels:
            row.append(value_getter(row_label, col_label))
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = number_format

    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 28
    for column_cells in sheet.iter_cols(min_col=2, max_col=len(col_labels) + 1):
        sheet.column_dimensions[column_cells[0].column_letter].width = 22


def write_tactic_sheet(
    sheet,
    row_tactics: list[str],
    col_ttps: list[str],
    data: CooccurrenceData,
    tactic_label: str,
) -> None:
    """Each cell: P(row_tactic | col_ttp) * P(col_ttp)"""
    header_font = Font(bold=True)
    priors = {ttp: prior_probability(data, ttp) for ttp in col_ttps}

    rows: list[tuple[str, float, list[float]]] = []  # (tactic, avg, [values])
    for tactic in row_tactics:
        values = [
            compute_tactic_given_ttp_probability(data, tactic, ttp) * priors[ttp]
            for ttp in col_ttps
        ]
        avg = sum(values) / len(values) if values else 0.0
        rows.append((tactic, avg, values))

    rows.sort(key=lambda r: r[1], reverse=True)

    col_labels = [f"{t}: {data.ttp_names[t]}" for t in col_ttps]
    sheet.freeze_panes = "C2"
    sheet.append([f"Tactic \\ {tactic_label} TTPs", "Weighted Average", *col_labels])
    for cell in sheet[1]:
        cell.font = header_font

    for tactic, avg, values in rows:
        sheet.append([tactic, avg, *values])

    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.0000"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 22
    for column_cells in sheet.iter_cols(min_col=3, max_col=len(col_ttps) + 2):
        sheet.column_dimensions[column_cells[0].column_letter].width = 22


def write_workbook(output_path: Path, data: CooccurrenceData) -> None:
    workbook = Workbook()
    all_ttp_ids = ordered_ttps(data.ttp_names)

    individual_sheet = workbook.active
    assert individual_sheet is not None
    individual_sheet.title = "individual"

    ttp_labels = [f"{t}: {data.ttp_names[t]}" for t in all_ttp_ids]
    write_matrix(
        sheet=individual_sheet,
        row_labels=ttp_labels,
        col_labels=ttp_labels,
        value_getter=lambda a_label, b_label: compute_individual_probability(
            data,
            a_label.split(":")[0].strip(),
            b_label.split(":")[0].strip(),
        ),
        number_format="0.0000",
        row_header="TTP A \\ TTP B",
    )

    ttps_by_tactic: dict[str, list[str]] = defaultdict(list)
    for ttp_id in all_ttp_ids:
        tactic = data.ttp_tactic.get(ttp_id)
        if tactic:
            ttps_by_tactic[tactic].append(ttp_id)

    sorted_tactics = sorted(ttps_by_tactic)

    for tactic in sorted_tactics:
        sheet = workbook.create_sheet(title=sheet_name_for_tactic(tactic))
        write_tactic_sheet(
            sheet=sheet,
            row_tactics=sorted_tactics,
            col_ttps=ttps_by_tactic[tactic],
            data=data,
            tactic_label=tactic,
        )

    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    dataset_path = Path(args.dataset).expanduser().resolve()
    if not dataset_path.exists():
        raise FileNotFoundError(f"Dataset not found: {dataset_path}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else dataset_path.with_name(f"{dataset_path.stem}-cooccurrence.xlsx")
    )

    reports = load_reports(dataset_path)
    data = build_cooccurrence_data(reports)
    write_workbook(output_path, data)

    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
