"""
Usage example:
    uv run ttp_next_tactic.py dataset.json \
        --group "Spearphishing TTPs:T1566,T1566.001,T1566.002,T1566.003,T1566.004" \
        --group "File and Directory Discovery:T1083" \
        --group "Data from Network Shared Drive:T1039" \
        --group "Data from Local System:T1005"
"""
import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ttp_iter import TTPRecord, load_reports, iter_consecutive_pairs, normalize_id

END_OF_CHAIN = "END OF CHAIN"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compute the next-tactic probabilities for specified TTP groups",
    )
    parser.add_argument("dataset", help="Path to the input dataset JSON file")
    parser.add_argument(
        "-o", "--output", default=None,
        help="Path to the output .xlsx file. Defaults to <dataset stem>-next-tactic.xlsx",
    )
    parser.add_argument(
        "--group",
        action="append",
        dest="groups",
        metavar="LABEL:TTP_ID[,TTP_ID,...]",
        help=(
            "A named group of TTPs to use as one sheet tab. "
            "Format: 'Sheet Label:T1234,T1234.001,...'. Repeat for multiple tabs."
        ),
    )
    return parser.parse_args()


def parse_groups(raw: list[str]) -> list[tuple[str, set[str]]]:
    groups: list[tuple[str, set[str]]] = []
    for entry in raw:
        if ":" not in entry:
            raise ValueError(f"--group must be 'Label:TTP_ID,...', got: {entry!r}")
        label, ids_str = entry.split(":", 1)
        ids = {normalize_id(t.strip()) for t in ids_str.split(",") if t.strip()}
        if not ids:
            raise ValueError(f"--group {entry!r} has no TTP IDs after the colon")
        groups.append((label.strip(), ids))
    return groups


def build_transition_counts(
    reports: list[dict[str, object]],
    all_ttp_ids: set[str],
) -> tuple[dict[str, dict[str, int]], dict[str, TTPRecord]]:
    counts: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    ttp_records: dict[str, TTPRecord] = {}

    for current, nxt in iter_consecutive_pairs(reports):
        ttp_records.setdefault(current.id, current)
        if nxt is not None:
            ttp_records.setdefault(nxt.id, nxt)

        if current.id not in all_ttp_ids:
            continue

        if nxt is None:
            counts[current.id][END_OF_CHAIN] += 1
        elif nxt.tactic:
            counts[current.id][nxt.tactic] += 1

    return counts, ttp_records


def aggregate_group_counts(
    group_ttp_ids: set[str],
    counts: dict[str, dict[str, int]],
) -> dict[str, int]:
    aggregated: dict[str, int] = defaultdict(int)
    for ttp_id in group_ttp_ids:
        for row, n in counts.get(ttp_id, {}).items():
            aggregated[row] += n
    return aggregated


def _record_label(record: TTPRecord) -> str:
    tactic = f" ({record.tactic})" if record.tactic else ""
    return f"{record.id}: {record.name}{tactic}"


def group_corner_label(
    group_ttp_ids: set[str],
    ttp_records: dict[str, TTPRecord],
) -> str:
    if len(group_ttp_ids) == 1:
        ttp_id = next(iter(group_ttp_ids))
        record = ttp_records.get(ttp_id)
        if record is not None:
            return _record_label(record)
        return ttp_id

    parts = [
        _record_label(ttp_records[ttp_id]) if ttp_id in ttp_records else ttp_id
        for ttp_id in sorted(group_ttp_ids)
    ]
    return ", ".join(parts)


def write_sheet(
    sheet,
    group_ttp_ids: set[str],
    ttp_records: dict[str, TTPRecord],
    counts: dict[str, dict[str, int]],
) -> None:
    aggregated = aggregate_group_counts(group_ttp_ids, counts)
    total = sum(aggregated.values())

    ordered_rows = sorted(aggregated, key=lambda row: -aggregated[row])

    corner = group_corner_label(group_ttp_ids, ttp_records)
    header_font = Font(bold=True)
    sheet.freeze_panes = "B2"
    sheet.append([corner, "Proportion", "Count"])
    for cell in sheet[1]:
        cell.font = header_font

    for row in ordered_rows:
        n = aggregated[row]
        proportion = n / total if total > 0 else 0.0
        sheet.append([row, proportion, n])

    for ws_row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in ws_row:
            cell.number_format = "0.0000"

    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 10


def sheet_name(label: str) -> str:
    invalid = r"\/*?:[]{}"
    cleaned = "".join(c if c not in invalid else "_" for c in label)
    return cleaned[:31]


def main() -> None:
    args = parse_args()

    if not args.groups:
        raise SystemExit("Provide at least one --group argument")

    groups = parse_groups(args.groups)
    all_ttp_ids = {ttp_id for _, ids in groups for ttp_id in ids}

    dataset_path = Path(args.dataset).expanduser().resolve()
    if not dataset_path.exists():
        raise FileNotFoundError(f"Dataset not found: {dataset_path}")

    output_path = (
        Path(args.output).expanduser().resolve()
        if args.output
        else dataset_path.with_name(f"{dataset_path.stem}-next-tactic.xlsx")
    )

    reports = load_reports(dataset_path)
    counts, ttp_records = build_transition_counts(reports, all_ttp_ids)

    workbook = Workbook()
    if workbook.active is not None:
        workbook.remove(workbook.active)

    for label, group_ttp_ids in groups:
        ws = workbook.create_sheet(title=sheet_name(label))
        write_sheet(ws, group_ttp_ids, ttp_records, counts)

    workbook.save(output_path)
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
