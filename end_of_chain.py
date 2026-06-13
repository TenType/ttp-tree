import argparse
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from typing import Callable, TypeVar

from ttp_iter import iter_ttps

REASONABLE_TACTICS = ["impact", "exfiltration", "stealth", "persistence", "lateral-movement"]
REASONABLE_COLOR = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

T = TypeVar("T")
def write_rows(ws: Worksheet, counter: Counter[T], col1: Callable[[T], str], col2: Callable[[T], str], get_tactic: Callable[[T], str]):
    total = counter.total()
    for key, count in counter.most_common():
        ws.append([col1(key), col2(key), count / total, count])
        ws[f"C{ws.max_row}"].number_format = "0.00%"
        if get_tactic(key) in REASONABLE_TACTICS:
            for cell in ws[ws.max_row]:
                cell.fill = REASONABLE_COLOR

def main():
    parser = argparse.ArgumentParser(
        description="Compute statistics for TTPs at the end of chains",
    )
    parser.add_argument("dataset", help="Path to the input dataset JSON file")
    args = parser.parse_args()

    dataset_path = Path(args.dataset)

    n = 0
    tactic_counter = Counter()
    technique_counter = Counter()

    for report in iter_ttps(dataset_path):
        tactic_counter[report.ttps[-1].tactic] += 1
        technique_counter[report.ttps[-1]] += 1
        n += 1

    total = tactic_counter.total()

    wb = Workbook()
    ws = wb[wb.sheetnames[0]]

    # header
    ws.append(["Category", "", "Proportion", "Count"])
    ws.freeze_panes = "A2"
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 25

    print(f"{n} reports")

    # reasonable statistics
    reasonable_str = ", ".join(sorted(REASONABLE_TACTICS, key=lambda x: tactic_counter[x], reverse=True))
    reasonable_count = sum(count for key, count in tactic_counter.items() if key in REASONABLE_TACTICS)

    ws.append([reasonable_str, "", reasonable_count / total, reasonable_count])
    ws[f"C{ws.max_row}"].number_format = "0.00%"
    for cell in ws[ws.max_row]:
        cell.fill = REASONABLE_COLOR

    unreasonable_tactics = [k for k in tactic_counter if k not in REASONABLE_TACTICS]
    unreasonable_str = ", ".join(sorted(unreasonable_tactics, key=lambda x: tactic_counter[x], reverse=True))
    unreasonable_count = total - reasonable_count

    ws.append([unreasonable_str, "", unreasonable_count / total, unreasonable_count])
    ws[f"C{ws.max_row}"].number_format = "0.00%"

    # tactic statistics
    ws.append([])
    write_rows(ws, tactic_counter, lambda x: x, lambda _: "", lambda x: x)

    # ttp statistics
    ws.append([])
    write_rows(ws, technique_counter, lambda ttp: f"{ttp.id}: {ttp.name}", lambda ttp: ttp.tactic, lambda ttp: ttp.tactic)
    
    wb.save(dataset_path.with_name(f"{dataset_path.stem}-end-of-chain.xlsx"))

if __name__ == "__main__":
    main()
